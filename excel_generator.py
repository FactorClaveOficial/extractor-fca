"""
Generador de Excel con 4 pestañas estructuradas
Crea reportes con información de facturas CFDI
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any
from datetime import datetime


class ExcelGenerator:
    """Generador de reportes en Excel a partir de datos CFDI"""
    
    def __init__(self):
        self.workbook = openpyxl.Workbook()
        self.workbook.remove(self.workbook.active)
        
        # Estilos
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def generate_report(self, data_list: List[Dict[str, Any]], output_path: str) -> bool:
        """
        Genera un Excel con 4 pestañas
        
        Args:
            data_list: Lista de diccionarios con datos CFDI
            output_path: Ruta del archivo Excel a crear
            
        Returns:
            True si fue exitoso, False si hubo error
        """
        try:
            # Crea las 4 pestañas
            self._create_headers_sheet(data_list)
            self._create_concepts_sheet(data_list)
            self._create_taxes_sheet(data_list)
            self._create_stamps_sheet(data_list)
            
            # Guarda el archivo
            self.workbook.save(output_path)
            return True
        
        except Exception as e:
            print(f"Error generating Excel: {str(e)}")
            return False
    
    def _create_headers_sheet(self, data_list: List[Dict[str, Any]]):
        """Crea pestaña de encabezados (1 fila por factura)"""
        ws = self.workbook.create_sheet("Encabezados", 0)
        
        # Columnas
        columns = [
            'Folio', 'Serie', 'Fecha', 'RFC Emisor', 'Nombre Emisor',
            'RFC Receptor', 'Nombre Receptor', 'Tipo Comprobante',
            'Subtotal', 'Descuento', 'Total', 'Moneda', 'Régimen Fiscal', 'Uso CFDI'
        ]
        
        # Encabezados
        for col_num, column_title in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = column_title
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.border
        
        # Datos
        for row_num, data in enumerate(data_list, 2):
            header = data.get('header', {})
            
            values = [
                header.get('folio', ''),
                header.get('serie', ''),
                header.get('fecha', ''),
                header.get('emisor_rfc', ''),
                header.get('emisor_nombre', ''),
                header.get('receptor_rfc', ''),
                header.get('receptor_nombre', ''),
                header.get('tipo_comprobante', ''),
                header.get('subtotal', ''),
                header.get('descuento', ''),
                header.get('total', ''),
                header.get('moneda', ''),
                header.get('emisor_regimen_fiscal', ''),
                header.get('receptor_uso_cfdi', '')
            ]
            
            for col_num, value in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = self.border
                cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Ajusta ancho de columnas
        for col_num, column_title in enumerate(columns, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 18
    
    def _create_concepts_sheet(self, data_list: List[Dict[str, Any]]):
        """Crea pestaña de conceptos (líneas de factura)"""
        ws = self.workbook.create_sheet("Conceptos", 1)
        
        # Columnas
        columns = [
            'Folio', 'Serie', 'Cantidad', 'Unidad', 'Clave Producto',
            'Descripción', 'Valor Unitario', 'Importe', 'Descuento'
        ]
        
        # Encabezados
        for col_num, column_title in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = column_title
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.border
        
        # Datos
        row_num = 2
        for data in data_list:
            header = data.get('header', {})
            concepts = data.get('concepts', [])
            
            for concept in concepts:
                values = [
                    header.get('folio', ''),
                    header.get('serie', ''),
                    concept.get('cantidad', ''),
                    concept.get('unidad', ''),
                    concept.get('clave_prod_serv', ''),
                    concept.get('descripcion', ''),
                    concept.get('valor_unitario', ''),
                    concept.get('importe', ''),
                    concept.get('descuento', '')
                ]
                
                for col_num, value in enumerate(values, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.border = self.border
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                
                row_num += 1
        
        # Ajusta ancho de columnas
        for col_num in range(1, len(columns) + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 15
    
    def _create_taxes_sheet(self, data_list: List[Dict[str, Any]]):
        """Crea pestaña de impuestos (trasladados y retenciones)"""
        ws = self.workbook.create_sheet("Impuestos", 2)
        
        # Columnas
        columns = [
            'Folio', 'Serie', 'Tipo', 'Impuesto', 'Tasa', 'Importe'
        ]
        
        # Encabezados
        for col_num, column_title in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = column_title
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.border
        
        # Datos
        row_num = 2
        for data in data_list:
            header = data.get('header', {})
            taxes = data.get('taxes', {})
            
            # Traslados
            for traslado in taxes.get('trasladados', []):
                values = [
                    header.get('folio', ''),
                    header.get('serie', ''),
                    'Traslado',
                    traslado.get('impuesto', ''),
                    traslado.get('tasa', ''),
                    traslado.get('importe', '')
                ]
                
                for col_num, value in enumerate(values, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.border = self.border
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                
                row_num += 1
            
            # Retenciones
            for retencion in taxes.get('retenciones', []):
                values = [
                    header.get('folio', ''),
                    header.get('serie', ''),
                    'Retención',
                    retencion.get('impuesto', ''),
                    retencion.get('tasa', ''),
                    retencion.get('importe', '')
                ]
                
                for col_num, value in enumerate(values, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.border = self.border
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                
                row_num += 1
        
        # Ajusta ancho de columnas
        for col_num in range(1, len(columns) + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 15
    
    def _create_stamps_sheet(self, data_list: List[Dict[str, Any]]):
        """Crea pestaña de timbres TFD"""
        ws = self.workbook.create_sheet("Timbres TFD", 3)
        
        # Columnas
        columns = [
            'Folio', 'Serie', 'UUID', 'Fecha Timbrado', 'RFC Proveedor Certificación',
            'Número Certificado SAT', 'Número Certificado CSD', 'Leyenda'
        ]
        
        # Encabezados
        for col_num, column_title in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = column_title
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.border
        
        # Datos
        for row_num, data in enumerate(data_list, 2):
            header = data.get('header', {})
            stamp = data.get('stamp', {})
            
            values = [
                header.get('folio', ''),
                header.get('serie', ''),
                stamp.get('uuid', ''),
                stamp.get('fecha_timbrado', ''),
                stamp.get('rfc_prov_certif', ''),
                stamp.get('numero_certificado_sat', ''),
                stamp.get('numero_certificado_csd', ''),
                stamp.get('leyenda', '')
            ]
            
            for col_num, value in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = self.border
                cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Ajusta ancho de columnas
        for col_num in range(1, len(columns) + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 20
